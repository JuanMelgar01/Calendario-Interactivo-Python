from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook


class ReadError(Exception):
    pass


REQUIRED_COLUMNS = ["title", "description", "category", "start"]
OPTIONAL_COLUMNS = ["end"]
TEMPLATE_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def read_events_xlsx(path: str, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    excel_path = Path(path)
    if not excel_path.exists():
        raise ReadError(f"No existe el archivo Excel: {path}")

    try:
        workbook = load_workbook(filename=str(excel_path), data_only=True)
    except Exception as exc:
        raise ReadError(f"No se pudo abrir el Excel: {exc}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ReadError(
                f"La hoja '{sheet_name}' no existe. "
                f"Hojas disponibles: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    header = [_normalize_cell_value(cell.value).lower() for cell in worksheet[1]]
    if not header or all(column == "" for column in header):
        raise ReadError("La primera fila del Excel debe contener encabezados.")

    column_index = {name: header.index(name) for name in header if name}
    missing = [column for column in REQUIRED_COLUMNS if column not in column_index]
    if missing:
        raise ReadError(
            f"Faltan columnas requeridas en Excel: {missing}. "
            f"Columnas esperadas: {TEMPLATE_COLUMNS}"
        )

    def get_value(row: tuple[Any, ...], column: str) -> str:
        if column not in column_index:
            return ""
        index = column_index[column]
        cell_value = row[index] if index < len(row) else None
        return _normalize_cell_value(cell_value)

    events: List[Dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(value is None or str(value).strip() == "" for value in row):
            continue

        events.append(
            {
                "title": get_value(row, "title"),
                "description": get_value(row, "description"),
                "category": get_value(row, "category"),
                "start": get_value(row, "start"),
                "end": get_value(row, "end"),
            }
        )

    return events


def create_excel_template(path: str, sheet_name: str = "Eventos") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(TEMPLATE_COLUMNS)
    worksheet.append(
        [
            "Reunion con equipo",
            "Revisar avances semanales",
            "Trabajo",
            "2026-01-08T10:30",
            "2026-01-08T11:15",
        ]
    )
    worksheet.append(
        [
            "Semana de examenes",
            "Repasar temario y hacer simulacros",
            "Estudio",
            "2026-01-07",
            "2026-01-12",
        ]
    )

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
